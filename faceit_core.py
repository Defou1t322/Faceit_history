#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FACEIT Match History — ядро (core)
----------------------------------
Чистая логика работы с FACEIT Data API без привязки к интерфейсу: HTTP-клиент
с троттлингом и ретраем 429, разбор истории/матчей, статистика по никам и
определение пати, экспорт. Используется и десктоп-версией (tkinter), и
веб-версией (локальный сервер).
"""

import re
import csv
import time
from datetime import datetime, timezone

import requests

API_BASE = "https://open.faceit.com/data/v4"
REQUEST_TIMEOUT = 30

# Минимальный интервал между запросами к API (сек). Сдерживает per-second лимит
# FACEIT, чтобы реже ловить 429. ~0.1 ≈ 10 запросов/сек.
MIN_REQUEST_INTERVAL = 0.1
# Предохранитель для режима "по датам": максимум матчей на игрока, если диапазон
# дат окажется слишком широким.
DATE_MODE_MAX = 2000

# Сколько раз ник должен оказаться в ОДНОЙ команде с искомым игроком, чтобы
# считать связь вероятной пати (а не случайным совпадением в матчмейкинге).
PARTY_MIN_MATCHES = 3
PARTY_MIN_RATIO = 0.6     # доля "вместе" среди всех совместных матчей

# Если нужно зашить свой ключ в exe для раздачи другим (чтобы им не вводить),
# впишите его сюда перед сборкой. Иначе оставьте пустым — ключ вводится в окне.
DEFAULT_API_KEY = ""

GAMES = ["cs2", "csgo", "dota2", "valorant"]

PERIOD_LAST = "Последние N матчей"
PERIOD_DATES = "Даты (от–до)"
PERIOD_OFFSET = "Номера матчей (offset)"
PERIOD_MODES = [PERIOD_LAST, PERIOD_DATES, PERIOD_OFFSET]


class Cancelled(Exception):
    """Поднимается, когда пользователь нажал «Отмена» во время запроса/паузы."""


# ---------------------------------------------------------------------------
# HTTP-клиент: один на поиск. Throttle + ретрай на 429 + счётчик запросов.
# ---------------------------------------------------------------------------
class Client:
    def __init__(self, api_key, on_count=None, cancel=None,
                 min_interval=MIN_REQUEST_INTERVAL):
        self.session = requests.Session()
        self.session.headers.update(
            {"Authorization": f"Bearer {api_key}", "accept": "application/json"})
        self.on_count = on_count
        self.cancel = cancel
        self.min_interval = min_interval
        self._last = 0.0
        self.count = 0

    def _sleep(self, secs):
        end = time.monotonic() + secs
        while True:
            rem = end - time.monotonic()
            if rem <= 0:
                return
            if self.cancel is not None and self.cancel.is_set():
                raise Cancelled()
            time.sleep(min(0.05, rem))

    def get(self, url, params=None, allow_404=False):
        if self.cancel is not None and self.cancel.is_set():
            raise Cancelled()
        r = None
        for attempt in range(6):
            # throttle: не чаще, чем раз в min_interval
            wait = self.min_interval - (time.monotonic() - self._last)
            if wait > 0:
                self._sleep(wait)
            self._last = time.monotonic()

            r = self.session.get(url, params=params, timeout=REQUEST_TIMEOUT)
            self.count += 1
            if self.on_count:
                self.on_count(self.count)

            if r.status_code == 429:
                ra = r.headers.get("Retry-After", "")
                delay = float(ra) if ra.replace(".", "", 1).isdigit() \
                    else min(2 ** attempt, 16)
                self._sleep(max(delay, 1.0))
                continue
            if allow_404 and r.status_code == 404:
                return None
            r.raise_for_status()
            return r.json()
        raise requests.HTTPError(
            "429 — лимит запросов FACEIT не отпускает даже после ретраев.",
            response=r)


# ---------------------------------------------------------------------------
# API-логика (через Client — никаких обращений к GUI)
# ---------------------------------------------------------------------------
def nickname_from_input(value: str) -> str:
    value = value.strip()
    m = re.search(r"/players/([^/?#]+)", value)
    return m.group(1) if m else value


def get_player_id(client: Client, nickname: str) -> str:
    data = client.get(f"{API_BASE}/players",
                      params={"nickname": nickname}, allow_404=True)
    if data is None:
        raise ValueError(f"Игрок '{nickname}' не найден на FACEIT.")
    return data["player_id"]


def get_history(client: Client, player_id: str, game: str, period: dict) -> list:
    """Возвращает список items истории согласно выбранному периоду.

    period:
      {"mode": "last",   "limit": N}
      {"mode": "offset", "start": S, "end": E}     # матчи [S, E)
      {"mode": "dates",  "from": ts|None, "to": ts|None, "max": M}
    """
    url = f"{API_BASE}/players/{player_id}/history"
    mode = period["mode"]
    items = []

    if mode == "last":
        limit, offset = period["limit"], 0
        while len(items) < limit:
            batch = min(100, limit - len(items))
            page = client.get(url, params={"game": game, "offset": offset,
                                           "limit": batch}).get("items", [])
            if not page:
                break
            items.extend(page)
            offset += batch
        return items[:limit]

    if mode == "offset":
        offset, end = period["start"], period["end"]
        while offset < end:
            batch = min(100, end - offset)
            page = client.get(url, params={"game": game, "offset": offset,
                                           "limit": batch}).get("items", [])
            if not page:
                break
            items.extend(page)
            offset += batch
        return items

    if mode == "dates":
        cap, offset = period.get("max", DATE_MODE_MAX), 0
        while len(items) < cap:
            batch = min(100, cap - len(items))
            params = {"game": game, "offset": offset, "limit": batch}
            if period.get("from"):
                params["from"] = period["from"]
            if period.get("to"):
                params["to"] = period["to"]
            page = client.get(url, params=params).get("items", [])
            if not page:
                break
            items.extend(page)
            offset += batch
        return items

    raise ValueError(f"Неизвестный режим периода: {mode}")


def parse_match(client: Client, match_id: str):
    data = client.get(f"{API_BASE}/matches/{match_id}/stats", allow_404=True)
    if not data or not data.get("rounds"):
        return None
    rnd = data["rounds"][0]
    rs = rnd.get("round_stats", {})
    teams = []
    for team in rnd.get("teams", []):
        tstats = team.get("team_stats", {})
        players = []
        for p in team.get("players", []):
            st = p.get("player_stats", {})
            k, d, a = st.get("Kills", "0"), st.get("Deaths", "0"), st.get("Assists", "0")
            players.append({
                "nickname": p.get("nickname"),
                "player_id": p.get("player_id"),
                "kills": k, "deaths": d, "assists": a,
                "kda": f"{k}/{d}/{a}",
                "kd_ratio": st.get("K/D Ratio", ""),
            })
        teams.append({
            "name": tstats.get("Team", team.get("team_id", "")),
            "won": tstats.get("Team Win") == "1",
            "final_score": tstats.get("Final Score", ""),
            "players": players,
        })
    return {
        "match_id": match_id,
        "map": rs.get("Map", ""),
        "score": rs.get("Score", "?"),
        "teams": teams,
    }


# ---------------------------------------------------------------------------
# Статистика по никам: частота + средний KDA (по всем игрокам матчей)
# ---------------------------------------------------------------------------
def _to_int(x) -> int:
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return 0


def _searched_team_indices(match: dict, searched_ids) -> set:
    """Индексы команд матча, в которых есть хотя бы один искомый игрок."""
    idx = set()
    for ti, team in enumerate(match["teams"]):
        if any(p.get("player_id") in searched_ids for p in team["players"]):
            idx.add(ti)
    return idx


def classify_role(d: dict) -> str:
    if d["is_searched"]:
        return "★ искомый"
    together = d["with"] + d["vs"]
    if (d["with"] >= PARTY_MIN_MATCHES and d["with"] >= d["vs"]
            and together and d["with"] / together >= PARTY_MIN_RATIO):
        return "пати?"
    if d["with"] > 0:
        return "вместе"
    return "против"


def compute_stats(results: list, searched_ids=frozenset()) -> list:
    """По всем игрокам всех матчей: частота, средний KDA и связь с искомыми.

    Для каждого ника считается, сколько раз он был в ОДНОЙ команде с искомым
    игроком ("with") и сколько раз — против ("vs"). На этом строится роль:
    искомый / пати? / вместе / против.

    Возвращает список словарей: искомые первыми, затем по частоте и "вместе".
    """
    agg = {}
    for m in results:
        s_teams = _searched_team_indices(m, searched_ids)
        any_searched = bool(s_teams)
        for ti, team in enumerate(m["teams"]):
            same = ti in s_teams
            for p in team["players"]:
                nick = p["nickname"]
                if not nick:
                    continue
                d = agg.setdefault(nick, {
                    "matches": 0, "k": 0, "d": 0, "a": 0, "kd_sum": 0.0,
                    "kd_n": 0, "with": 0, "vs": 0, "is_searched": False,
                    "player_id": p.get("player_id"),
                })
                d["matches"] += 1
                d["k"] += _to_int(p["kills"])
                d["d"] += _to_int(p["deaths"])
                d["a"] += _to_int(p["assists"])
                try:
                    d["kd_sum"] += float(p["kd_ratio"])
                    d["kd_n"] += 1
                except (TypeError, ValueError):
                    pass
                if p.get("player_id") in searched_ids:
                    d["is_searched"] = True
                elif any_searched:
                    if same:
                        d["with"] += 1
                    else:
                        d["vs"] += 1

    rows = []
    for nick, d in agg.items():
        n = d["matches"]
        rows.append({
            "nickname": nick,
            "matches": n,
            "avg_k": d["k"] / n,
            "avg_d": d["d"] / n,
            "avg_a": d["a"] / n,
            "avg_kda": f"{d['k'] / n:.1f}/{d['d'] / n:.1f}/{d['a'] / n:.1f}",
            "avg_kd": d["kd_sum"] / d["kd_n"] if d["kd_n"] else 0.0,
            "with": d["with"],
            "vs": d["vs"],
            "is_searched": d["is_searched"],
            "role": classify_role(d),
            "player_id": d["player_id"],
        })
    rows.sort(key=lambda r: (r["is_searched"], r["matches"], r["with"]),
              reverse=True)
    return rows


def party_nicknames(stats: list) -> set:
    """Ники, помеченные как вероятная пати (часто в одной команде с искомым)."""
    return {s["nickname"] for s in stats if s["role"] == "пати?"}


STATS_COLUMNS = ["nickname", "role", "matches", "with_searched", "vs_searched",
                 "avg_kills", "avg_deaths", "avg_assists", "avg_kda", "avg_kd_ratio"]


def stats_rows_flat(stats: list) -> list:
    return [[s["nickname"], s["role"], s["matches"], s["with"], s["vs"],
             f"{s['avg_k']:.2f}", f"{s['avg_d']:.2f}", f"{s['avg_a']:.2f}",
             s["avg_kda"], f"{s['avg_kd']:.2f}"] for s in stats]


def export_stats_csv(path: str, stats: list):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(STATS_COLUMNS)
        w.writerows(stats_rows_flat(stats))


def stats_to_tsv(stats: list) -> str:
    lines = ["\t".join(STATS_COLUMNS)]
    for row in stats_rows_flat(stats):
        lines.append("\t".join(str(x) for x in row))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Экспорт матчей: CSV / Excel / TSV (для Google Sheets)
# ---------------------------------------------------------------------------
EXPORT_COLUMNS = ["match_id", "date", "map", "score", "team", "result",
                  "nickname", "kills", "deaths", "assists", "kda", "kd_ratio",
                  "searched", "relation", "party"]


def unix_to_str(ts) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d %H:%M")
    except (ValueError, OSError, OverflowError):
        return ""


def parse_date_to_unix(text: str, end_of_day=False):
    """'YYYY-MM-DD' -> unix (UTC). Пусто -> None. Ошибка -> ValueError."""
    text = text.strip()
    if not text:
        return None
    dt = datetime.strptime(text, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    ts = int(dt.timestamp())
    return ts + 86399 if end_of_day else ts


def flatten_rows(results: list, searched_ids=frozenset(), party_nicks=frozenset()) -> list:
    """Разворачивает результаты в плоские строки: 1 строка = игрок в матче.

    Доп. колонки: searched (искомый игрок), relation (self/team/enemy —
    относительно искомых в этом матче), party (◆ если ник — вероятная пати).
    """
    rows = []
    for m in results:
        s_teams = _searched_team_indices(m, searched_ids)
        for ti, team in enumerate(m["teams"]):
            res = "win" if team["won"] else "lose"
            for p in team["players"]:
                is_s = p.get("player_id") in searched_ids
                if is_s:
                    relation = "self"
                elif ti in s_teams:
                    relation = "team"
                elif s_teams:
                    relation = "enemy"
                else:
                    relation = ""
                rows.append([
                    m["match_id"], m.get("date", ""), m.get("map", ""), m["score"],
                    team["name"], res, p["nickname"], p["kills"], p["deaths"],
                    p["assists"], p["kda"], p["kd_ratio"],
                    "yes" if is_s else "",
                    relation,
                    "party?" if (not is_s and p["nickname"] in party_nicks) else "",
                ])
    return rows


def export_csv(path: str, results: list, searched_ids=frozenset(), party_nicks=frozenset()):
    # utf-8-sig (BOM) + ';' — чтобы Excel в RU/UA-локали корректно показал
    # кириллицу и сам разбил по столбцам.
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(EXPORT_COLUMNS)
        w.writerows(flatten_rows(results, searched_ids, party_nicks))


def export_xlsx(path: str, results: list, searched_ids=frozenset(), party_nicks=frozenset()):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Для экспорта в Excel нужен модуль openpyxl "
                           "(pip install openpyxl). Используйте CSV или TSV.")
    wb = Workbook()
    ws = wb.active
    ws.title = "FACEIT history"
    ws.append(EXPORT_COLUMNS)
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = h_font
        c.fill = h_fill
        c.alignment = Alignment(horizontal="center")
    win_fill = PatternFill("solid", fgColor="C6EFCE")
    lose_fill = PatternFill("solid", fgColor="F2F2F2")
    self_font = Font(bold=True, color="0A4DA0")     # искомый игрок
    party_font = Font(bold=True, color="7A1FA2")    # вероятная пати
    for row in flatten_rows(results, searched_ids, party_nicks):
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=6)  # столбец result
        cell.fill = win_fill if row[5] == "win" else lose_fill
        nick_cell = ws.cell(row=ws.max_row, column=7)  # столбец nickname
        if row[12] == "yes":        # searched
            nick_cell.font = self_font
        elif row[14] == "party?":   # party
            nick_cell.font = party_font
    for i, w in enumerate([38, 16, 14, 10, 18, 8, 22, 7, 7, 8, 10, 9,
                           9, 9, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def results_to_tsv(results: list, searched_ids=frozenset(), party_nicks=frozenset()) -> str:
    lines = ["\t".join(EXPORT_COLUMNS)]
    for row in flatten_rows(results, searched_ids, party_nicks):
        lines.append("\t".join(str(x) for x in row))
    return "\n".join(lines)


def http_error_message(status) -> str:
    return {
        400: "400 — некорректный запрос (проверьте ник/игру/период).",
        401: "401 — неверный или отсутствующий API-ключ.",
        403: "403 — доступ запрещён (проверьте права ключа).",
        404: "404 — не найдено.",
        429: "429 — превышен лимит запросов. Подождите и повторите.",
        503: "503 — сервис FACEIT временно недоступен.",
    }.get(status, f"Ошибка HTTP {status}.")

